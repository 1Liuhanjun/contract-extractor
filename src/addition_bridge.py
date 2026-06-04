"""Bridge to the original Addition appendix project.

The Addition project is treated as a black box after OCR. We run an Addition
adapter in a subprocess, feed it the shared OCR markdown, and read the generated
"邮政" sheet for display/merge. This avoids importing Addition's absolute module
names (schemas/layers/providers) into the main app process and keeps both
extraction workflows isolated.
"""

import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
ADDITION_DIR = PROJECT_DIR / "Addition"
ADDITION_OUTPUT_DIR = ADDITION_DIR / "output" / "contracts"
POSTAL_SHEET_PREFIX = "邮政"


def run_addition_appendix(source_path, markdown_path=None):
    """Run Addition's original post-OCR pipeline and return appendix rows.

    The original Addition project starts from PDF OCR. In this integrated app,
    OCR/text extraction is shared by the main project, so Addition receives a
    markdown-like text file and runs its original post-OCR business pipeline.
    """
    source = Path(source_path)
    if not ADDITION_DIR.exists():
        return _error(f"未找到 Addition 项目目录: {ADDITION_DIR}")
    if not markdown_path:
        return _error("缺少共享 OCR markdown，无法以第二档模式运行 Addition 附表流程。")
    markdown = Path(markdown_path)
    if not markdown.exists():
        return _error(f"共享 OCR markdown 不存在: {markdown}")

    ADDITION_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = ADDITION_OUTPUT_DIR / f"合同台账_{source.stem}.xlsx"
    cmd = [
        sys.executable,
        "shared_ocr_runner.py",
        "--source",
        str(source.resolve()),
        "--markdown",
        str(markdown.resolve()),
        "--output",
        str(xlsx_path.resolve()),
    ]
    try:
        completed = subprocess.run(
            cmd,
            cwd=str(ADDITION_DIR),
            capture_output=True,
            text=True,
            timeout=2400,
            check=False,
        )
    except Exception as exc:
        return _error(f"调用原版附表项目失败: {exc}")

    if completed.returncode != 0:
        details = (completed.stderr or completed.stdout or "").strip()[-3000:]
        return _error(f"Addition 共享 OCR 流程运行失败(exit={completed.returncode}): {details}")

    runner_result = _parse_runner_result(completed.stdout)
    if runner_result.get("xlsx"):
        xlsx_path = Path(runner_result["xlsx"])
    if not xlsx_path.exists():
        return _error(f"Addition 共享 OCR 流程未生成 Excel: {xlsx_path}")

    try:
        rows = _read_postal_rows(xlsx_path)
    except Exception as exc:
        return _error(f"读取原版附表 Excel 失败: {exc}")

    return {
        "success": True,
        "found": bool(rows),
        "row_count": len(rows),
        "rows": rows,
        "source": "Addition(shared_text)",
        "source_xlsx": str(xlsx_path),
        "runner_result": runner_result,
        "warnings": [] if rows else ["原版附表项目运行成功，但邮政 sheet 没有数据行。"],
    }


def _parse_runner_result(stdout):
    marker = "SHARED_OCR_RESULT_JSON="
    for line in reversed((stdout or "").splitlines()):
        if line.startswith(marker):
            try:
                return json.loads(line[len(marker):])
            except json.JSONDecodeError:
                return {}
    return {}


def _read_postal_rows(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    ws = next((sheet for sheet in wb.worksheets if sheet.title.startswith(POSTAL_SHEET_PREFIX)), None)
    if ws is None:
        wb.close()
        raise ValueError("未找到邮政 sheet")

    headers = []
    for cell in ws[1]:
        header = str(cell.value).strip() if cell.value is not None else ""
        headers.append(header)

    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in excel_row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = excel_row[idx] if idx < len(excel_row) else ""
            item[header] = "" if value is None else value
        rows.append(item)
    wb.close()
    return rows


def _error(message):
    return {
        "success": False,
        "found": False,
        "row_count": 0,
        "rows": [],
        "source": "Addition",
        "error": message,
        "warnings": [message],
    }
