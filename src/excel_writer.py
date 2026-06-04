"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

Excel写入器
"""
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MAIN_SHEET_NAME = "主表"
LEGACY_MAIN_SHEET_NAME = "合同信息提取"
APPENDIX_SHEET_NAME = "附表"
APPENDIX_HEADERS = [
    "省", "市", "合同编码", "业务类型", "分包号", "线路名称", "邮路性质", "里程",
    "2.75吨/3吨", "5吨", "4.2米", "8吨", "12吨/9.6米", "15吨", "20吨/12.5",
    "25吨", "30吨、17.5", "40吨A", "元/趟/条/40吨B", "有效期开始时间",
    "有效期结束时间", "公司主体", "类型", "备注", "是否退线", "退线时间",
    "合同编号", "返程线路编码", "延期结束时间",
]


def _styles():
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_fill, header_font, thin_border


def _ensure_sheet(wb, title, create=True):
    if title in wb.sheetnames:
        return wb[title]
    if title == MAIN_SHEET_NAME and LEGACY_MAIN_SHEET_NAME in wb.sheetnames:
        ws = wb[LEGACY_MAIN_SHEET_NAME]
        ws.title = MAIN_SHEET_NAME
        return ws
    if title == MAIN_SHEET_NAME and wb.sheetnames:
        ws = wb.active
        if ws.max_row <= 1:
            ws.title = MAIN_SHEET_NAME
            return ws
    return wb.create_sheet(title) if create else None


def _write_headers(ws, headers):
    header_fill, header_font, thin_border = _styles()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"


def _write_main_rows(ws, standardized_results, metadata_list=None, start_row=2):
    from field_knowledge_base import EXCEL_HEADERS

    _, _, thin_border = _styles()
    for row_offset, std_result in enumerate(standardized_results):
        row_num = start_row + row_offset
        metadata = metadata_list[row_offset] if metadata_list else {}

        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            if header in std_result and std_result[header] not in (None, "", "null"):
                value = std_result[header]
            elif header in metadata and metadata[header] not in (None, "", "null"):
                value = metadata[header]
            else:
                value = ""

            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(len(str(value)) > 20))
            cell.border = thin_border
        ws.row_dimensions[row_num].height = 45


def _appendix_rows_from_results(appendix_results_list):
    rows = []
    for appendix in appendix_results_list or []:
        if not isinstance(appendix, dict):
            continue
        for row in appendix.get("rows", []) or []:
            if isinstance(row, dict):
                rows.append(row)
    return rows


def _write_appendix_rows(ws, appendix_rows, start_row=2):
    _, _, thin_border = _styles()
    for row_offset, row_data in enumerate(appendix_rows):
        row_num = start_row + row_offset
        for col_idx, header in enumerate(APPENDIX_HEADERS, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_num, column=col_idx, value="" if value in (None, "null") else value)
            cell.alignment = Alignment(vertical="center", wrap_text=(len(str(value)) > 20))
            cell.border = thin_border
        ws.row_dimensions[row_num].height = 36


def write_results(standardized_results, output_path, metadata_list=None, appendix_results_list=None):
    """
    将标准化结果写入Excel

    参数:
        standardized_results: list[dict]，每个dict是标准字段值
        output_path: 输出文件路径
        metadata_list: list[dict]，内部字段的值（登记日期/快递单号等）
    """
    from field_knowledge_base import EXCEL_HEADERS, HEADER_TO_KEY

    wb = Workbook()
    ws = wb.active
    ws.title = MAIN_SHEET_NAME

    _write_headers(ws, EXCEL_HEADERS)
    _write_main_rows(ws, standardized_results, metadata_list=metadata_list, start_row=2)

    # 行高
    col_widths = {
        "登记日期": 12, "项目名称/编号": 18, "客户分类": 10, "业务类型": 10,
        "合同类型": 10, "省": 6, "市": 6, "客户名称": 35,
        "合同主体": 30, "合同编码": 22, "合同名称": 30,
        "合同开始时间": 14, "合同结束时间": 14, "是否完成签订": 12,
        "是否同步财务": 12, "合同预警提醒": 12, "账期（天）": 10,
        "保证金（万元）": 12, "税率": 10,
        "是否有旺季补偿": 12, "旺季补偿时间": 55, "旺季补偿规则": 50,
        "补偿比例": 10, "是否有油价联动": 12, "油价基准(元/升）": 12,
        "是否有疫情补贴": 12, "补贴标准": 20,
        "联系人": 10, "电话": 14, "地址": 40,
        "快递单号": 16, "钉钉审批单号": 22
    }
    for idx, header in enumerate(EXCEL_HEADERS, 1):
        width = col_widths.get(header, 12)
        ws.column_dimensions[get_column_letter(idx)].width = width

    appendix_rows = _appendix_rows_from_results(appendix_results_list)
    appendix_ws = wb.create_sheet(APPENDIX_SHEET_NAME)
    _write_headers(appendix_ws, APPENDIX_HEADERS)
    _write_appendix_rows(appendix_ws, appendix_rows, start_row=2)
    for idx, header in enumerate(APPENDIX_HEADERS, 1):
        width = 16
        if header in ("线路名称", "公司主体"):
            width = 28
        elif header in ("有效期开始时间", "有效期结束时间", "合同编码"):
            width = 18
        appendix_ws.column_dimensions[get_column_letter(idx)].width = width


    wb.save(output_path)
    count = len(standardized_results)
    print(f"\n  [完成] 已写入 {count} 条记录到: {output_path}")
    return output_path


def append_results(output_path, standardized_results, metadata_list=None, appendix_results_list=None):
    """Append main rows and appendix rows to an existing combined workbook."""
    from field_knowledge_base import EXCEL_HEADERS

    if not os.path.exists(output_path):
        return write_results(standardized_results, output_path, metadata_list, appendix_results_list)

    wb = load_workbook(output_path)
    main_ws = _ensure_sheet(wb, MAIN_SHEET_NAME)
    if main_ws.max_row < 1 or not main_ws.cell(row=1, column=1).value:
        _write_headers(main_ws, EXCEL_HEADERS)
    start_row = main_ws.max_row + 1
    _write_main_rows(main_ws, standardized_results, metadata_list=metadata_list, start_row=start_row)

    appendix_ws = _ensure_sheet(wb, APPENDIX_SHEET_NAME)
    if appendix_ws.max_row < 1 or not appendix_ws.cell(row=1, column=1).value:
        _write_headers(appendix_ws, APPENDIX_HEADERS)
    appendix_rows = _appendix_rows_from_results(appendix_results_list)
    if appendix_rows:
        _write_appendix_rows(appendix_ws, appendix_rows, start_row=appendix_ws.max_row + 1)

    wb.save(output_path)
    print(f"\n  [完成] 已追加 {len(standardized_results)} 条主表记录到: {output_path}")
    return output_path
