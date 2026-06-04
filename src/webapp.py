#!/usr/bin/env python3
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

合同提取 Web 应用
================
基于 Python 标准库的轻量 Web 服务器。
无需 Flask，python3 即可运行。

启动：python3 webapp.py
访问：http://localhost:8080
"""
import os
import sys
import json
import io
import re
import mimetypes
import traceback
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse
from datetime import datetime

# PROJECT_DIR = 项目根目录（src/ 的上级）
PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_DIR)
sys.path.insert(0, os.path.join(PROJECT_DIR, 'src'))

# 从 .env 文件加载 API Key
try:
    from dotenv import load_dotenv
    env_path = os.path.join(PROJECT_DIR, '.env')
    if os.path.exists(env_path):
        load_dotenv(env_path)
except ImportError:
    pass

UPLOAD_DIR = os.path.join(PROJECT_DIR, "data/uploads")
REVIEWS_DIR = os.path.join(PROJECT_DIR, "data/reviews")
RESULTS_DIR = os.path.join(PROJECT_DIR, "data/results")
OCR_TEXT_DIR = os.path.join(PROJECT_DIR, "data/ocr_texts")
LEARNING_DIR = os.path.join(PROJECT_DIR, "data/learning_candidates")
CUMULATIVE_EXCEL = os.path.join(RESULTS_DIR, "合同提取汇总.xlsx")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(REVIEWS_DIR, exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)
os.makedirs(OCR_TEXT_DIR, exist_ok=True)
os.makedirs(LEARNING_DIR, exist_ok=True)

CONTRACT_CODE_COUNTERS = {}


# ============================================================
# 核心处理逻辑
# ============================================================
def _contract_subject_code(subject_name):
    """合同主体 → 编码前缀。未知主体保守归为 QTZT，避免误写成 BHWL。"""
    text = str(subject_name or "")
    rules = [
        ("BHWL", ["北京博华物流有限公司", "博华物流", "北京博华"]),
        ("TJZZ", ["天津智猪网网络科技有限公司", "智猪网", "天津智猪"]),
    ]
    for code, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return code
    return "QTZT"


def _customer_code(customer_type, customer_name):
    """客户分类/名称 → 合同编码中段。"""
    text = f"{customer_type or ''}{customer_name or ''}"
    if "邮政" in text or "中国邮政" in text:
        return "YZ邮政"
    if "京东" in text:
        return "JD京东"
    return "QT其他"


def _next_contract_code_sequence(code_prefix, today_str):
    """根据累积 Excel 和本轮进程内计数生成当日序号。"""
    cache_key = f"{code_prefix}-{today_str}"
    max_seq = CONTRACT_CODE_COUNTERS.get(cache_key, 0)

    if os.path.exists(CUMULATIVE_EXCEL):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(CUMULATIVE_EXCEL, read_only=True, data_only=True)
            ws = wb.active
            header_row = [cell.value for cell in ws[1]]
            if "合同编码" in header_row:
                code_col = header_row.index("合同编码") + 1
                prefix = f"{code_prefix}-{today_str}"
                for row in range(2, ws.max_row + 1):
                    value = ws.cell(row=row, column=code_col).value
                    if not value:
                        continue
                    value = str(value).strip()
                    if not value.startswith(prefix):
                        continue
                    m = re.search(r"(\d{3})$", value)
                    if m:
                        max_seq = max(max_seq, int(m.group(1)))
            wb.close()
        except Exception:
            pass

    next_seq = max_seq + 1
    CONTRACT_CODE_COUNTERS[cache_key] = next_seq
    return str(next_seq).zfill(3)


def _parse_date_value(value):
    """解析常见合同日期格式，返回 date；解析失败返回 None。"""
    if value is None:
        return None
    if hasattr(value, "date") and hasattr(value, "strftime"):
        try:
            return value.date()
        except Exception:
            pass
    if hasattr(value, "strftime"):
        try:
            return value
        except Exception:
            pass

    text = str(value).strip()
    if not text or text in ("null", "None", "/"):
        return None

    m = re.search(r"(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})", text)
    if not m:
        m = re.search(r"(\d{4})(\d{2})(\d{2})", text)
    if not m:
        return None

    try:
        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
    except Exception:
        return None


def _calculate_warning_days(end_date_value):
    """合同预警提醒：合同结束时间 - 今天，过期返回负数。"""
    end_date = _parse_date_value(end_date_value)
    if end_date is None:
        return ""
    from datetime import date
    return str((end_date - date.today()).days)


def process_contract(file_path, api_key, api_provider="anthropic"):
    from llm_client import LLMClient
    from stage1_fact_extraction import extract_facts
    from stage2_standardizer import standardize_fields
    from few_shot_examples import get_all_examples

    load_meta = {}
    text = _load_text(file_path, metadata=load_meta)
    if text is None or len(text) < 50:
        return {"error": "文件内容过短或无法读取"}

    llm = LLMClient(provider=api_provider, api_key=api_key)
    facts = extract_facts(text, llm)
    if facts is None:
        return {"error": "Stage 1 事实抽取失败"}

    # Python级安全校验：中标人信息必须包含乙方公司名，否则标记为不可靠
    # 这个操作 LLM 做不好（容易把甲方描述误判为乙方）
    their_name = ""
    bizhong_fact = facts.get("乙方_全称", {})
    if isinstance(bizhong_fact, dict):
        their_name = bizhong_fact.get("value", "") or "北京博华物流有限公司"
    if not their_name:
        their_name = "北京博华物流有限公司"

    for fact_key in ["中标人信息", "中标人", "成交人信息"]:
        fv = facts.get(fact_key, "")
        if isinstance(fv, dict):
            fv = fv.get("value", "")
        fv = str(fv)
        if fv and fv != "未提及" and their_name not in fv:
            # 中标描述不含乙方公司名 → 不可靠，标注
            facts[fact_key] = {"value": "未提及（原描述不含乙方公司名，已过滤）",
                              "evidence": fv[:200],
                              "_filtered": True}

    examples = get_all_examples()
    standardized = standardize_fields(facts, examples, llm)
    if standardized is None:
        return {"error": "Stage 2 标准化映射失败"}

    # Python级修正：如果合同类型=主选/备选，且没有有效的乙方侧中标依据，
    # 回退只看线路表维度
    if standardized.get("合同类型") == "主选/备选":
        has_valid_bidding = False
        for fk in ["中标人信息", "中标人", "成交人信息"]:
            fv = facts.get(fk, "")
            if isinstance(fv, dict):
                fv_data = fv
            else:
                fv_data = {"value": str(fv)}
            if not fv_data.get("_filtered") and fv_data.get("value", "") not in ("未提及", ""):
                has_valid_bidding = True
                break
        if not has_valid_bidding:
            # 没有有效的中标人信息，只看线路表
            line_info = facts.get("线路表格", facts.get("线路信息", ""))
            line_str = str(line_info) if line_info else ""
            has_zhu = any(kw in line_str for kw in ["正式", "主选"])
            has_bei = any(kw in line_str for kw in ["备用", "备选"])
            if has_zhu and has_bei:
                standardized["合同类型"] = "主选/备选"
            elif has_bei:
                standardized["合同类型"] = "备选"
            else:
                standardized["合同类型"] = "主选"

    # 自动生成字段（2026-06-02 业务确认）
    import re as _re
    from datetime import date as _date

    # 登记日期 = 今天
    standardized["登记日期"] = _date.today().strftime("%Y/%m/%d")

    # 项目名称/编号 = 不填
    standardized["项目名称/编号"] = ""

    # 合同主体 = Stage 2 从合同提取（乙方名称，可能是博华或其他公司）
    # 不再硬编码，由 LLM 从合同原文中提取

    # 合同编码 = 合同主体缩写-客户缩写-YYYYMMDDNNN
    today_str = _date.today().strftime("%Y%m%d")
    subject_code = _contract_subject_code(standardized.get("合同主体", ""))
    customer_code = _customer_code(standardized.get("客户分类", ""), standardized.get("客户名称", ""))
    code_prefix = f"{subject_code}-{customer_code}"
    seq = _next_contract_code_sequence(code_prefix, today_str)
    standardized["合同编码"] = f"{code_prefix}-{today_str}{seq}"

    # 合同名称 = 第一条线路名 + "一干运输合同"
    route_name = ""
    # 从 facts 或 contract_text 中找第一条线路
    for key in ["第一条线路", "线路名称", "首条线路"]:
        f = facts.get(key, "")
        if isinstance(f, dict):
            f = f.get("value", "")
        if f and f != "未提及":
            route_name = str(f).strip()
            break
    if not route_name:
        # fallback: 从合同文本中搜索线路表
        m = _re.search(r'(\w+)-(\w+)</td>', text)
        if m:
            route_name = m.group(0).replace("</td>", "").strip()
    if not route_name:
        route_name = "未命名线路"
    # 清洗线路名（去掉HTML残留）
    route_name = _re.sub(r'<[^>]+>', '', route_name).strip()
    standardized["合同名称"] = f"{route_name}一干运输合同"

    # 合同预警提醒 = 结束时间 - 今天
    end_str = str(standardized.get("合同结束时间", "")).strip()
    standardized["合同预警提醒"] = _calculate_warning_days(end_str)

    # 是否完成签订 / 是否同步财务 = 不填
    standardized["是否完成签订"] = ""
    standardized["是否同步财务"] = ""

    # 联系人/电话/地址/快递单号/钉钉审批单号 = 不填
    for f in ["联系人", "电话", "地址", "快递单号", "钉钉审批单号"]:
        standardized[f] = ""

    # 适用 "/" 风格的空值字段（答案习惯填 "/" 表示无）
    slash_fields = ["保证金（万元）", "是否有疫情补贴", "补贴标准",
                    "是否有旺季补偿", "旺季补偿时间", "旺季补偿规则", "补偿比例",
                    "是否有油价联动", "油价基准(元/升）", "账期（天）"]
    for f in slash_fields:
        if f in standardized and standardized[f] in (None, "", "null", "None"):
            standardized[f] = "/"

    response = {"facts": facts, "result": standardized, "file_name": os.path.basename(file_path), "contract_text": text}
    if load_meta.get("ocr"):
        response["ocr"] = load_meta["ocr"]
    return response


def _load_text(file_path, metadata=None):
    if file_path.lower().endswith('.pdf'):
        try:
            from ocr_client import PaddleOCRClient
            client = PaddleOCRClient()
            print(f"  [OCR] PDF 文件，开始调用 PaddleOCR: {os.path.basename(file_path)}")
            ocr_result = client.extract_text_from_file(file_path, save_dir=OCR_TEXT_DIR)
            if metadata is not None:
                metadata["ocr"] = {
                    "job_id": ocr_result.get("job_id"),
                    "model": ocr_result.get("model"),
                    "saved_text_path": ocr_result.get("saved_text_path"),
                    "saved_jsonl_path": ocr_result.get("saved_jsonl_path"),
                }
            return ocr_result.get("text", "")
        except Exception as e:
            print(f"  [OCR错误] PDF 转文本失败: {e}")
            return None

    if file_path.endswith('.docx'):
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            with zipfile.ZipFile(file_path) as z:
                xml_content = z.read('word/document.xml')
                root = ET.fromstring(xml_content)
                texts = []
                ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
                for p in root.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
                    para = []
                    for t in p.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                        if t.text and not t.text.strip().startswith('<') and not t.text.strip().startswith('http'):
                            para.append(t.text)
                    if para:
                        texts.append(''.join(para))
                return '\n'.join(texts)
        except Exception:
            return None
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            return f.read()
    except Exception:
        return None


def _append_to_cumulative(standardized_list):
    """将提取结果追加到累积 Excel，不覆盖已有行"""
    from excel_writer import write_results
    from field_knowledge_base import EXCEL_HEADERS

    if not os.path.exists(CUMULATIVE_EXCEL):
        write_results(standardized_list, CUMULATIVE_EXCEL)
        print(f"  📊 创建累积 Excel: {CUMULATIVE_EXCEL}")
        return

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wb = load_workbook(CUMULATIVE_EXCEL)
    ws = wb.active
    start_row = ws.max_row + 1

    for idx, std in enumerate(standardized_list):
        row_num = start_row + idx
        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            raw = std.get(header, "")
            val = "" if raw in (None, "", "null") else str(raw)
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(len(val) > 20))
            cell.border = thin_border

    wb.save(CUMULATIVE_EXCEL)
    print(f"  📊 已追加 {len(standardized_list)} 行到累积 Excel (共 {start_row + len(standardized_list) - 2} 行数据)")


def _is_numeric_equivalent(a, b):
    """判断两个值是否在数值上等价（处理 % 与小数互转、精度差异）"""
    def to_num(v):
        v = v.replace(" ", "").replace(",", "")
        if v.endswith("%"):
            num_part = v[:-1].strip()
        else:
            num_part = v
        # 必须是纯数字才能转
        if not num_part.replace(".", "").replace("-", "").isdigit():
            return None
        try:
            n = float(num_part)
            # 如果值小于1，可能是小数表示的百分比（0.09 → 9）
            if n < 1:
                return n * 100
            return n
        except ValueError:
            return None

    na = to_num(a)
    nb = to_num(b)
    if na is None or nb is None:
        return False
    # 差异在5%以内算一致
    if na == 0 and nb == 0:
        return True
    max_val = max(abs(na), abs(nb))
    if max_val == 0:
        return True
    return abs(na - nb) / max_val < 0.05


def verify_results(extracted, reference_data):
    """将提取结果与参考答案对比（验证32列中除8个明确不填字段外的24个字段）

    判定规则：
    - "/" 和空字符串视为等价（都表示"无此信息"）
    - "YYYY/MM/DD" 和 "YYYY-MM-DD" 视为等价
    - 语义等价算一致（如："否"≈"/"、数值精度差异等）
    """
    from field_knowledge_base import get_verifiable_headers

    valid_headers = get_verifiable_headers()

    def normalize(v, field=None, row_data=None):
        """标准化：None/空/null→''，'/'→''，去空格"""
        if v is None or v == "null":
            return ""
        v = str(v).strip()
        if v == "/":
            return ""
        if field == "合同预警提醒":
            if v.startswith("=") or "DATEDIF" in v.upper() or "NOW()" in v.upper():
                calculated = _calculate_warning_days((row_data or {}).get("合同结束时间", ""))
                if calculated:
                    return calculated
            calculated = _calculate_warning_days(v)
            if calculated:
                return calculated
        m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", v)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        return v

    # 长文本字段（允许子串匹配）
    FUZZY_FIELDS = {"客户名称", "旺季补偿时间", "旺季补偿规则", "旺季补偿", "合同名称"}

    def _fuzzy_match(a, b, min_overlap=0.6):
        """模糊匹配：双向包含 或 公共子串占比超过阈值"""
        if not a or not b:
            return False
        # 直接包含
        if a in b or b in a:
            return True
        # 计算最长公共子串占比
        shorter = a if len(a) <= len(b) else b
        longer = b if len(a) <= len(b) else a
        # 找最长公共子串
        max_len = 0
        for i in range(len(shorter)):
            for j in range(i + 1, len(shorter) + 1):
                if shorter[i:j] in longer:
                    max_len = max(max_len, j - i)
                else:
                    break
        return max_len / len(shorter) >= min_overlap

    comparisons = []
    correct_count = 0
    wrong_count = 0

    for header in valid_headers:
        ai_val = normalize(extracted.get(header, ""), header, extracted)
        ref_val = normalize(reference_data.get(header, ""), header, reference_data)

        # 两个都为空 → 一致
        if ai_val == "" and ref_val == "":
            status = "一致（均为空）"
            correct_count += 1
        # 完全一致
        elif ai_val == ref_val:
            status = "✅ 一致"
            correct_count += 1
        # 仅空格差异
        elif ai_val.replace(" ", "") == ref_val.replace(" ", ""):
            status = "✅ 一致（仅空格差异）"
            correct_count += 1
        # 语义等价：AI="否" vs 参考="/"（都表示没有）
        elif ai_val in ("否", "无") and ref_val == "":
            status = "✅ 一致（语义等价：无此信息）"
            correct_count += 1
        elif ai_val == "" and ref_val in ("否", "无"):
            status = "✅ 一致（语义等价：无此信息）"
            correct_count += 1
        # 数值精度差异（如 10% vs 10.00%, 317.842 vs 317.8, 9% vs 0.09）
        elif _is_numeric_equivalent(ai_val, ref_val):
            status = "✅ 一致（数值等价）"
            correct_count += 1
        # AI有值，参考为空
        elif ai_val != "" and ref_val == "":
            status = "⚠️ AI多填"
            wrong_count += 1
        # AI为空，参考有值
        elif ai_val == "" and ref_val != "":
            status = "⚠️ AI未提取到"
            wrong_count += 1
        # 长文本模糊匹配（客户名称、旺季等字段允许子串匹配）
        elif header in FUZZY_FIELDS and _fuzzy_match(ai_val, ref_val):
            status = "✅ 一致（模糊匹配）"
            correct_count += 1
        else:
            status = "❌ 不一致"
            wrong_count += 1

        comparisons.append({
            "field": header,
            "ai_value": ai_val if ai_val else "（空）",
            "ref_value": ref_val if ref_val else "（空）",
            "status": status,
        })

    return {
        "comparisons": comparisons,
        "summary": {
            "total": len(comparisons),
            "correct": correct_count,
            "wrong": wrong_count,
        }
    }


def generate_review(contract_name, extracted, reference_data, comparison):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r'[^\w\-_]', '_', contract_name)

    s = comparison["summary"]
    accuracy = s['correct'] / s['total'] * 100 if s['total'] > 0 else 0

    lines = []
    lines.append(f"# 合同提取验证报告\n")
    lines.append(f"- **合同**: {contract_name}")
    lines.append(f"- **时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- **总字段**: {s['total']} | **一致**: {s['correct']} | **不一致**: {s['wrong']} | **准确率**: {accuracy:.1f}%\n")

    lines.append(f"## 逐字段对比\n")
    lines.append(f"| 字段 | AI提取值 | 参考答案 | 判定 |")
    lines.append(f"|------|---------|---------|------|")
    for c in comparison["comparisons"]:
        ai_v = c['ai_value'][:50] if c['ai_value'] else "（空）"
        ref_v = c['ref_value'][:50] if c['ref_value'] else "（空）"
        lines.append(f"| {c['field']} | {ai_v} | {ref_v} | {c['status']} |")

    wrongs = [c for c in comparison["comparisons"] if "不一致" in c['status'] or "多填" in c['status'] or "未提取" in c['status']]
    if wrongs:
        lines.append(f"\n## 差异明细\n")
        for w in wrongs:
            lines.append(f"- **{w['field']}**: AI=`{w['ai_value']}` vs 参考=`{w['ref_value']}`")

    content = "\n".join(lines)
    # 描述性文件名
    prov = extracted.get("省", "")
    city = extracted.get("市", "")
    label = f"{prov}{city}" if prov and city else safe_name
    review_path = os.path.join(REVIEWS_DIR, f"{timestamp}_{label}_验证报告.md")
    with open(review_path, 'w', encoding='utf-8') as f:
        f.write(content)

    json_data = {
        "contract_name": contract_name,
        "timestamp": timestamp,
        "extracted": {k: str(v) for k, v in extracted.items()},
        "reference": {k: str(v) for k, v in reference_data.items()},
        "comparison": comparison,
    }
    json_path = os.path.join(REVIEWS_DIR, f"{timestamp}_{safe_name}.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

    return review_path, json_path


# ============================================================
# Multipart 解析器（处理 FormData 上传）
# ============================================================
def parse_multipart(body, content_type):
    """解析 multipart/form-data，返回文件列表"""
    boundary = None
    for part in content_type.split(";"):
        part = part.strip()
        if part.startswith("boundary="):
            boundary = part[9:]
            break
    if not boundary:
        return []

    boundary = boundary.encode()
    parts = body.split(b"--" + boundary)
    files = []

    for part in parts:
        if part == b"" or part == b"--\r\n" or part == b"\r\n" or part == b"-\r\n":
            continue

        # 分离 header 和 body
        header_end = part.find(b"\r\n\r\n")
        if header_end == -1:
            continue

        header_bytes = part[:header_end]
        file_body = part[header_end + 4:]

        # 去掉尾部\r\n
        while file_body.endswith(b"\r\n"):
            file_body = file_body[:-2]
        while file_body.endswith(b"\r\n"):
            file_body = file_body[:-2]

        header_str = header_bytes.decode("utf-8", errors="replace")

        # 提取文件名
        filename = None
        m = re.search(r'filename="([^"]*)"', header_str)
        if m:
            filename = m.group(1)

        if not filename:
            continue

        files.append({"filename": filename, "content": file_body})

    return files


# ============================================================
# HTTP 服务器
# ============================================================
CACHED_HTML = None


class ContractHandler(BaseHTTPRequestHandler):

    def _send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        if length == 0:
            return b""
        return self.rfile.read(length)

    # ---- GET ----
    def do_GET(self):
        path = urlparse(self.path).path
        if path == "/" or path == "/index.html":
            self._serve_html()
        elif path == "/api/files":
            self._list_files()
        elif path == "/api/learning-candidates":
            self._list_learning_candidates()
        elif path.startswith("/static/"):
            self._serve_static()
        else:
            self._send_json({"error": "Not found"}, 404)

    def _serve_html(self):
        global CACHED_HTML
        if CACHED_HTML is None:
            html_path = os.path.join(PROJECT_DIR, "web", "templates", "index.html")
            if os.path.exists(html_path):
                with open(html_path, "r", encoding="utf-8") as f:
                    CACHED_HTML = f.read()
            else:
                CACHED_HTML = "<html><body><h1>templates/index.html 未找到</h1></body></html>"
        body = CACHED_HTML.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _serve_static(self):
        path = urlparse(self.path).path  # e.g. /static/style.css
        filename = path[len("/static/"):]
        filepath = os.path.join(PROJECT_DIR, "web", "static", filename)
        if os.path.exists(filepath):
            mime, _ = mimetypes.guess_type(filepath)
            with open(filepath, "rb") as f:
                body = f.read()
            self.send_response(200)
            self.send_header("Content-Type", mime or "application/octet-stream")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        else:
            self._send_json({"error": f"File not found: {filename}"}, 404)

    def _list_files(self):
        files = [f for f in os.listdir(UPLOAD_DIR) if f.lower().endswith((".txt", ".docx", ".pdf"))]
        self._send_json({"files": files})

    def _list_learning_candidates(self):
        """列出旁路候选记忆库。只读展示，不影响正式抽取。"""
        records = []
        for filename in sorted(os.listdir(LEARNING_DIR), reverse=True):
            if not filename.endswith(".json"):
                continue
            path = os.path.join(LEARNING_DIR, filename)
            if not os.path.isfile(path):
                continue
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                continue

            candidates = data.get("candidates", [])
            if not isinstance(candidates, list):
                candidates = []
            records.append({
                "file_name": filename,
                "created_at": datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S"),
                "agent_role": data.get("agent_role", "learning_companion"),
                "summary": data.get("summary", {}),
                "candidates": candidates,
            })

        self._send_json({
            "success": True,
            "records": records,
            "count": len(records),
        })

    # ---- POST ----
    def do_POST(self):
        path = urlparse(self.path).path
        try:
            if path == "/api/upload":
                self._handle_upload()
            elif path == "/api/process":
                self._handle_process()
            elif path == "/api/export":
                self._handle_export()
            elif path == "/api/verify":
                self._handle_verify()
            elif path == "/api/generate-review":
                self._handle_generate_review()
            elif path == "/api/clear":
                self._handle_clear()
            elif path == "/api/ref-data":
                self._handle_ref_data()
            elif path == "/api/llm-review":
                self._handle_llm_review()
            elif path == "/api/learning-agent":
                self._handle_learning_agent()
            elif path == "/api/learning-candidate-status":
                self._handle_learning_candidate_status()
            elif path == "/api/chat":
                self._handle_chat()
            elif path == "/api/shutdown":
                self._send_json({"success": True, "message": "服务器已关闭"})
                import threading
                threading.Timer(0.5, lambda: os._exit(0)).start()
            else:
                self._send_json({"error": "Not found"}, 404)
        except Exception as e:
            traceback.print_exc()
            self._send_json({"error": str(e)}, 500)

    def _handle_upload(self):
        """上传文件（FormData）—— 先清空旧的再上传，确保只有一份"""
        body = self._read_body()
        ct = self.headers.get("Content-Type", "")

        # 清空旧的临时文件
        for f in os.listdir(UPLOAD_DIR):
            fpath = os.path.join(UPLOAD_DIR, f)
            if os.path.isfile(fpath):
                os.remove(fpath)

        files = parse_multipart(body, ct)
        saved = []
        for f in files:
            if not f["filename"].lower().endswith((".txt", ".docx", ".pdf")):
                continue
            save_path = os.path.join(UPLOAD_DIR, f["filename"])
            with open(save_path, "wb") as out:
                out.write(f["content"])
            saved.append(f["filename"])

        self._send_json({"success": True, "files": saved})

    def _handle_process(self):
        """处理已上传的文件"""
        body = self._read_body()
        try:
            data = json.loads(body)
        except Exception:
            data = {}

        api_provider = data.get("api_provider", "anthropic")

        # 优先取前端传入的key，否则按provider从环境变量取
        env_map = {
            "anthropic": "ANTHROPIC_API_KEY",
            "deepseek": "DEEPSEEK_API_KEY",
            "deepseek-v4-pro": "DEEPSEEK_API_KEY",
            "openai": "OPENAI_API_KEY",
        }
        api_key = data.get("api_key", "")
        if not api_key:
            env_key = env_map.get(api_provider, "ANTHROPIC_API_KEY")
            api_key = os.environ.get(env_key, "")
            if not api_key:
                self._send_json({"error": f"API Key 未设置，请在 .env 中填写 {env_key} 或在前端输入"}, 400)
                return

        file_list = data.get("files", os.listdir(UPLOAD_DIR))
        file_list = [f for f in file_list if f.lower().endswith((".txt", ".docx", ".pdf"))]

        if not file_list:
            self._send_json({"error": "没有可处理的文件"}, 400)
            return

        results = []
        for fname in file_list:
            fpath = os.path.join(UPLOAD_DIR, fname)
            result = process_contract(fpath, api_key, api_provider)
            result["file_name"] = fname
            results.append(result)

        # 自动追加到累积 Excel
        standardized_list = [r["result"] for r in results if "result" in r and r["result"]]
        if standardized_list:
            _append_to_cumulative(standardized_list)

        self._send_json({"success": True, "results": results})

    def _handle_export(self):
        """导出 Excel"""
        body = self._read_body()
        try:
            data = json.loads(body)
        except Exception:
            data = {}

        results = data.get("results", [])
        standardized_list = [r["result"] for r in results if "result" in r and r["result"]]

        if not standardized_list:
            self._send_json({"error": "没有有效的结果"}, 400)
            return

        try:
            from excel_writer import write_results
            output_path = os.path.join(PROJECT_DIR, "data", "_temp_export.xlsx")
            write_results(standardized_list, output_path)

            with open(output_path, "rb") as f:
                excel_data = f.read()
            os.remove(output_path)

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            # HTTP 头只接受 ASCII，中文文件名需要编码
            import urllib.parse
            safe_filename = "contract_results.xlsx"
            disposition = f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{urllib.parse.quote('合同提取结果.xlsx', safe='')}"
            self.send_header("Content-Disposition", disposition)
            self.send_header("Content-Length", str(len(excel_data)))
            self.end_headers()
            self.wfile.write(excel_data)
        except Exception as e:
            self._send_json({"error": f"导出失败: {e}"}, 500)

    def _handle_ref_data(self):
        """
        上传参考答案 Excel，解析所有行数据
        前端先调用这个接口上传参考Excel，拿到解析后的数据
        """
        body = self._read_body()
        ct = self.headers.get("Content-Type", "")

        files = parse_multipart(body, ct)
        xlsx_file = None
        for f in files:
            if f["filename"].endswith(".xlsx"):
                xlsx_file = f
                break

        if not xlsx_file:
            self._send_json({"error": "未找到Excel文件"}, 400)

        from openpyxl import load_workbook
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(xlsx_file["content"])
        tmp.close()

        wb = load_workbook(tmp.name)
        # 读最后一个表（前两个是线路价格明细，"新合同台账"在最后）
        ws = wb[wb.sheetnames[-1]]
        headers = []
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col).value
            if h:
                # 去除表头前的 * 符号（如 *客户名称 → 客户名称）
                h = h.lstrip("*")
                headers.append(h)

        rows = []
        for row in range(2, ws.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                val = ws.cell(row=row, column=col).value
                key = header
                # 转换日期格式
                if hasattr(val, 'strftime'):
                    val = val.strftime("%Y-%m-%d")
                # 转换百分比
                elif isinstance(val, float) and 0 < val < 1 and '比例' in str(header):
                    val = f"{val*100:.2f}%"
                elif isinstance(val, float) and 0 < val < 1 and '税率' in str(header):
                    val = f"{val*100:.0f}%"
                row_data[key] = val
            rows.append(row_data)

        os.unlink(tmp.name)

        self._send_json({
            "success": True,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
        })

    def _handle_verify(self):
        """
        验证：前端发送 extract_results + ref_row
        用第2行（安庆）和第3行（成都）逐行验证
        """
        body = self._read_body()
        data = json.loads(body)
        extracted = data.get("extracted", {})
        ref_row = data.get("ref_row", {})
        row_index = data.get("row_index", 2)

        comparison = verify_results(extracted, ref_row)
        self._send_json({
            "success": True,
            "row_index": row_index,
            "comparison": comparison,
        })

    def _handle_generate_review(self):
        """生成复盘报告"""
        body = self._read_body()
        data = json.loads(body)

        contract_name = data.get("contract_name", "unknown")
        extracted = data.get("extracted", {})
        reference = data.get("reference", {})

        comparison = verify_results(extracted, reference)
        review_path, json_path = generate_review(contract_name, extracted, reference, comparison)

        self._send_json({
            "success": True,
            "review_path": review_path,
            "json_path": json_path,
            "comparison": comparison,
        })

    def _handle_clear(self):
        for f in os.listdir(UPLOAD_DIR):
            fpath = os.path.join(UPLOAD_DIR, f)
            if os.path.isfile(fpath):
                os.remove(fpath)
        self._send_json({"success": True})

    def _handle_chat(self):
        """LLM 答疑：根据事实抽取+映射结果+合同原文回答用户问题"""
        body = self._read_body()
        try:
            data = json.loads(body)
        except Exception:
            self._send_json({"error": "无效数据"}, 400)
            return

        question = data.get("question", "")
        facts = data.get("facts", {})
        standardized = data.get("standardized", {})
        contract_text = data.get("contract_text", "")

        if not question:
            self._send_json({"error": "请输入问题"}, 400)
            return

        api_key = data.get("api_key", "")
        api_provider = data.get("api_provider", "deepseek-v4-pro")
        if not api_key:
            env_map = {"deepseek-v4-pro": "DEEPSEEK_API_KEY", "deepseek": "DEEPSEEK_API_KEY",
                       "anthropic": "ANTHROPIC_API_KEY", "openai": "OPENAI_API_KEY"}
            api_key = os.environ.get(env_map.get(api_provider, "DEEPSEEK_API_KEY"), "")
        if not api_key:
            self._send_json({"error": "未找到API Key"}, 400)
            return

        from llm_client import LLMClient
        client = LLMClient(provider=api_provider, api_key=api_key)

        # 构建上下文
        facts_summary = json.dumps(facts, ensure_ascii=False, indent=2)[:3000]
        result_summary = json.dumps(standardized, ensure_ascii=False, indent=2)[:2000]
        contract_summary = (contract_text or "")[:4000]

        system = """你是一个合同提取系统的答疑助手。你可以看到：
1. Stage 1 事实抽取结果（从合同原文提取的事实+证据）
2. Stage 2 标准化映射结果（最终输出的字段值）
3. 合同原文片段

请根据以上信息回答用户关于字段提取的问题。如果某个字段不匹配，可以从事实抽取阶段和映射规则两个角度分析原因。回答简洁、具体、引用原文。"""

        user = f"## 事实抽取结果\n```json\n{facts_summary}\n```\n\n## 标准化结果\n```json\n{result_summary}\n```\n\n## 合同原文\n```\n{contract_summary}\n```\n\n## 用户问题\n{question}"

        result = client.call(system, user, max_tokens=2048)
        if result is None:
            self._send_json({"error": "LLM调用失败"}, 500)
            return

        # result 可能是 dict（JSON）或纯文本
        answer = result.get("answer", str(result)) if isinstance(result, dict) else str(result)
        self._send_json({"success": True, "answer": answer})

    def _handle_learning_agent(self):
        """学习精灵：生成候选经验，不修改正式业务规则"""
        body = self._read_body()
        try:
            data = json.loads(body)
        except Exception:
            self._send_json({"error": "无效数据"}, 400)
            return

        facts = data.get("facts", {})
        standardized = data.get("standardized", {})
        contract_text = data.get("contract_text", "")
        discrepancies = data.get("discrepancies", [])
        llm_review = data.get("llm_review", [])
        api_key = data.get("api_key", "")
        api_provider = data.get("api_provider", "deepseek-v4-pro")

        if not standardized:
            self._send_json({"error": "缺少标准化结果，请先完成提取"}, 400)
            return
        if not contract_text and not facts:
            self._send_json({"error": "缺少合同原文和事实抽取结果，请先完成提取"}, 400)
            return

        if not api_key:
            env_map = {
                "anthropic": "ANTHROPIC_API_KEY",
                "deepseek": "DEEPSEEK_API_KEY",
                "deepseek-v4-pro": "DEEPSEEK_API_KEY",
                "openai": "OPENAI_API_KEY",
            }
            api_key = os.environ.get(env_map.get(api_provider, "DEEPSEEK_API_KEY"), "")
        if not api_key:
            self._send_json({"error": "未找到API Key"}, 400)
            return

        from llm_client import LLMClient
        client = LLMClient(provider=api_provider, api_key=api_key)

        facts_summary = json.dumps(facts, ensure_ascii=False, indent=2)[:5000]
        result_summary = json.dumps(standardized, ensure_ascii=False, indent=2)[:3500]
        diff_summary = json.dumps(discrepancies, ensure_ascii=False, indent=2)[:2500] if discrepancies else "[]"
        review_summary = json.dumps(llm_review, ensure_ascii=False, indent=2)[:3500] if llm_review else "[]"
        if contract_text:
            contract_snippet = contract_text[:7000] + "\n...（中间省略）...\n" + contract_text[-4000:]
        else:
            contract_snippet = ""

        system_prompt = """你是“合同学习精灵 Agent”，职责是在正式业务流程旁边总结候选经验。

最高规则：
- 你不能修改正式业务规则。
- 你不能削弱或覆盖字段知识库。
- 你只能生成“候选学习项”，供人工审核。
- 候选学习项默认不参与提取，不进入 Stage 2 prompt，不写入 field_knowledge_base.py 或 few_shot_examples.py。
- 如果证据不足，不要提出规则更新。

输出必须是 JSON 对象，格式：
{
  "agent_role": "learning_companion",
  "summary": {
    "candidate_count": 0,
    "high_confidence": 0,
    "needs_business_review": 0
  },
  "candidates": [
    {
      "field": "字段名",
      "candidate_type": "rule_update | few_shot_example | caution_note | no_action",
      "observed_problem": "这次观察到的问题",
      "proposed_learning": "候选经验/候选规则/候选少样本描述",
      "evidence": "支持该候选经验的合同原文或复核依据",
      "source": "validation_diff | llm_review | facts | contract_text",
      "confidence": "high | medium | low",
      "target": "field_knowledge_base.py | few_shot_examples.py | docs/字段提取业务规则手册.md | none",
      "requires_human_approval": true,
      "status": "pending_review",
      "why_not_auto_apply": "为什么不能自动生效"
    }
  ]
}

候选学习项原则：
- 只从明确差异、LLM复核结论、合同原文证据中总结。
- 不要把一次偶然现象上升为通用规则；不确定就输出 no_action 或 low confidence。
- 参考答案可能有错，所以不能仅凭参考答案提出规则更新。
- 更推荐输出 caution_note 或 few_shot_example，而不是直接 rule_update。
- 所有 candidates 都必须 requires_human_approval=true 且 status=pending_review。
"""

        user_prompt = f"""## Stage 1 事实抽取结果
```json
{facts_summary}
```

## Stage 2 标准化结果
```json
{result_summary}
```

## 验证差异（如有）
```json
{diff_summary}
```

## LLM复核结果（如有）
```json
{review_summary}
```

## 合同原文摘要
```
{contract_snippet}
```

请只输出“候选学习项”，不要输出正式规则修改。"""

        agent_result = client.call(system_prompt, user_prompt, max_tokens=4096)
        if agent_result is None:
            self._send_json({"error": "学习精灵调用失败"}, 500)
            return

        if not isinstance(agent_result, dict):
            self._send_json({"error": "学习精灵返回格式异常"}, 500)
            return

        candidates = agent_result.get("candidates", [])
        if not isinstance(candidates, list):
            candidates = []
            agent_result["candidates"] = candidates
        for candidate in candidates:
            if isinstance(candidate, dict):
                candidate["requires_human_approval"] = True
                candidate["status"] = "pending_review"
        agent_result["agent_role"] = "learning_companion"
        agent_result["summary"] = {
            "candidate_count": len(candidates),
            "high_confidence": sum(1 for c in candidates if isinstance(c, dict) and c.get("confidence") == "high"),
            "needs_business_review": sum(1 for c in candidates if isinstance(c, dict) and c.get("requires_human_approval")),
        }

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prov = standardized.get("省", "")
        city = standardized.get("市", "")
        location = f"{prov}{city}" if prov and city else "合同"
        json_path = os.path.join(LEARNING_DIR, f"{timestamp}_{location}_学习候选.json")
        md_path = os.path.join(LEARNING_DIR, f"{timestamp}_{location}_学习候选.md")

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(agent_result, f, ensure_ascii=False, indent=2)

        lines = [
            "# 学习精灵候选经验",
            "",
            f"- **时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"- **统计**: {json.dumps(agent_result.get('summary', {}), ensure_ascii=False)}",
            f"- **状态**: 候选记录，仅供人工审核，不参与正式提取",
            "",
            "## 候选项",
            "",
        ]
        for item in candidates:
            if not isinstance(item, dict):
                continue
            lines.append(f"### {item.get('field', '')}")
            lines.append(f"- **类型**: {item.get('candidate_type', '')}")
            lines.append(f"- **置信度**: {item.get('confidence', '')}")
            lines.append(f"- **目标**: {item.get('target', '')}")
            lines.append(f"- **状态**: {item.get('status', 'pending_review')}")
            lines.append(f"- **观察问题**: {item.get('observed_problem', '')}")
            lines.append(f"- **候选经验**: {item.get('proposed_learning', '')}")
            lines.append(f"- **依据**: {item.get('evidence', '')}")
            lines.append(f"- **不自动生效原因**: {item.get('why_not_auto_apply', '')}")
            lines.append("")

        with open(md_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        self._send_json({
            "success": True,
            "learning": agent_result,
            "json_path": json_path,
            "md_path": md_path,
        })

    def _handle_learning_candidate_status(self):
        """人工标记候选记忆状态。确认也只表示“已确认候选”，不会自动生效。"""
        body = self._read_body()
        try:
            data = json.loads(body)
        except Exception:
            self._send_json({"error": "无效数据"}, 400)
            return

        file_name = os.path.basename(data.get("file_name", ""))
        status = data.get("status", "")
        review_note = data.get("review_note", "")
        allowed_status = {"pending_review", "approved", "rejected", "archived"}

        try:
            candidate_index = int(data.get("candidate_index", -1))
        except Exception:
            candidate_index = -1

        if not file_name.endswith(".json") or not file_name:
            self._send_json({"error": "候选文件名无效"}, 400)
            return
        if status not in allowed_status:
            self._send_json({"error": "候选状态无效"}, 400)
            return

        path = os.path.join(LEARNING_DIR, file_name)
        if not os.path.isfile(path):
            self._send_json({"error": "候选文件不存在"}, 404)
            return

        try:
            with open(path, "r", encoding="utf-8") as f:
                record = json.load(f)
        except Exception as e:
            self._send_json({"error": f"候选文件读取失败: {e}"}, 500)
            return

        candidates = record.get("candidates", [])
        if not isinstance(candidates, list) or candidate_index < 0 or candidate_index >= len(candidates):
            self._send_json({"error": "候选项序号无效"}, 400)
            return

        candidate = candidates[candidate_index]
        if not isinstance(candidate, dict):
            self._send_json({"error": "候选项格式无效"}, 400)
            return

        candidate["status"] = status
        candidate["review_note"] = review_note
        candidate["reviewed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        candidate["requires_human_approval"] = True
        candidate["applied_to_business_logic"] = False

        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(record, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self._send_json({"error": f"候选文件保存失败: {e}"}, 500)
            return

        self._send_json({
            "success": True,
            "file_name": file_name,
            "candidate_index": candidate_index,
            "candidate": candidate,
            "message": "候选状态已更新，未应用到正式业务逻辑",
        })

    def _handle_llm_review(self):
        """LLM复核：读取合同原文，逐字段验证 AI/参考答案哪个正确"""
        body = self._read_body()
        try:
            data = json.loads(body)
        except Exception:
            self._send_json({"error": "无效的请求数据"}, 400)
            return

        discrepancies = data.get("discrepancies", [])
        api_key = data.get("api_key", "")
        api_provider = data.get("api_provider", "deepseek-v4-pro")

        if not discrepancies:
            self._send_json({"error": "没有差异需要复核"}, 400)
            return

        # 取API Key
        if not api_key:
            env_map = {
                "anthropic": "ANTHROPIC_API_KEY",
                "deepseek": "DEEPSEEK_API_KEY",
                "deepseek-v4-pro": "DEEPSEEK_API_KEY",
                "openai": "OPENAI_API_KEY",
            }
            api_key = os.environ.get(env_map.get(api_provider, "DEEPSEEK_API_KEY"), "")
        if not api_key:
            self._send_json({"error": "未找到API Key"}, 400)
            return

        # 优先用前端传来的合同文本（从 processResults 取出），不再依赖文件
        contract_text = data.get("contract_text", "")
        if not contract_text:
            # fallback：从 _uploads/ 读取
            for fname in os.listdir(UPLOAD_DIR):
                if fname.lower().endswith((".txt", ".docx", ".pdf")):
                    try:
                        text = _load_text(os.path.join(UPLOAD_DIR, fname))
                        if text and len(text) > 100:
                            contract_text = text
                            break
                    except Exception:
                        continue
        if not contract_text:
            self._send_json({"error": "未找到合同原文，请重新上传并提取一次"}, 400)
            return

        contract_snippet = contract_text[:6000] + "\n...（中间省略）...\n" + contract_text[-3000:]

        # 获取合同地区信息用于文件名
        standardized = data.get("standardized", {})
        prov = standardized.get("省", "")
        city = standardized.get("市", "")

        # 获取 Stage 1 事实抽取结果（让复核 LLM 看到 AI 已经提取了什么）
        extracted_facts = data.get("extracted_facts", {})

        # Python级验证：中标的描述必须包含乙方公司名，否则在传给复核 LLM 前过滤掉
        their_name = ""
        bizhong_raw = extracted_facts.get("乙方_全称", {})
        if isinstance(bizhong_raw, dict):
            their_name = bizhong_raw.get("value", "") or "北京博华物流有限公司"
        if not their_name:
            their_name = "北京博华物流有限公司"

        facts_summary = ""
        if extracted_facts:
            facts_summary = "\n## AI (Stage 1) 事实抽取结果\n以下是从合同原文中提取的事实（含原文位置），复核时参考：\n\n"
            for key, val in extracted_facts.items():
                if isinstance(val, dict):
                    v = val.get("value", "")
                    ev = val.get("evidence", "")
                    # 中标人相关字段 → 验证主语
                    if "中标" in key or "成交人" in key:
                        if v and v != "未提及" and their_name not in str(v):
                            facts_summary += f"- **{key}**: ⚠️ [已过滤] 原描述不含乙方名({their_name})，不是乙方身份描述。应丢弃。原描述: {str(v)[:100]}\n"
                            continue
                    facts_summary += f"- **{key}**: {v}\n"
                    if ev:
                        facts_summary += f"  证据: {ev}\n"
                else:
                    facts_summary += f"- **{key}**: {val}\n"

        from llm_client import LLMClient
        client = LLMClient(provider=api_provider, api_key=api_key)

        diff_text = ""
        for i, d in enumerate(discrepancies, 1):
            diff_text += f"差异{i}:\n  字段: {d['field']}\n  AI提取值: {d.get('ai_value', '（空）')}\n  参考答案: {d.get('ref_value', '（空）')}\n\n"

        system_prompt = """你是一个合同审核专家。请根据合同原文和AI事实抽取结果，对每个差异字段做严格复核。

**⚠️ 合同类型复核的强制规则（最高优先级）**：
复核"合同类型"差异时，你引用的中标/成交描述中**必须包含乙方公司全称**（如"北京博华物流有限公司"）。如果一段描述的主语只是甲方（如"江苏省分公司""中国邮政"），不含乙方公司名，则该描述与乙方类型**完全无关**，不能作为判断依据。
- ✅ 有效："北京博华物流有限公司为...第一成交人"
- ❌ 无效："江苏省分公司...第一成交人（主供应商）"——不含乙方名，Ignore

**对其他差异字段的复核要求**：

对每个差异字段，你需要在合同原文中搜索相关信息，然后输出标准化JSON。

**代码生成字段的复核要求**：
- 登记日期：按系统当天日期生成，格式可为 YYYY/MM/DD 或 YYYY-MM-DD。
- 合同编码：按“合同主体缩写-客户缩写-YYYYMMDDNNN”生成；北京博华物流有限公司→BHWL，天津智猪网网络科技有限公司→TJZZ，邮政→YZ邮政，京东→JD京东。
- 合同名称：按第一条线路名 + “一干运输合同”生成。
- 合同预警提醒：按合同结束时间 - 系统当天日期计算剩余天数。
- 这些字段合同原文中通常没有最终填写值，复核时应检查生成规则和依赖字段，而不是要求合同原文直接出现最终值。

{
  "field": "字段名",
  "ai_value": "AI提取的值",
  "ref_value": "参考答案的值",
  "judgment": "AI正确 | 参考答案正确 | 都不正确",
  "is_field_missing": true或false,
  "contract_text": "合同原文中对应的句子（直接引用）",
  "contract_location": "该句在合同中的位置（如：第四条第二款、第十五条第一项等）",
  "explanation": "复核依据：合同写了XXX，因此XXX"
}

要求：
- contract_text 必须是从合同原文中逐字引用的完整句子
- contract_location 必须指明在第几条、第几款、第几项
- is_field_missing 表示该字段在合同中是否根本不存在
- judgment 要根据合同原文与AI值、参考值的比较来判定谁正确
- 如果合同没有，is_field_missing=true，contract_text和contract_location填空

必须引用合同原文，不能凭空猜测。输出JSON数组。"""

        user_prompt = f"## 合同原文（首尾摘要）\n---\n{contract_snippet}\n---\n\n{facts_summary}\n## 需要复核的差异\n{diff_text}\n\n请逐条分析。"

        result = client.call(system_prompt, user_prompt, max_tokens=8192)
        if result is None:
            self._send_json({"error": "LLM复核调用失败"}, 500)
            return

        # 保存复核档案
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        review_lines = []
        review_lines.append("# LLM复核档案\n")
        review_lines.append(f"- **时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        review_lines.append(f"- **复核字段数**: {len(discrepancies)}\n")
        review_lines.append(f"## 复核结果\n")

        judgments = {"AI正确": 0, "参考答案正确": 0, "都不正确": 0}
        for item in result:
            j = item.get("judgment", "")
            key = j if j in judgments else ("都不正确" if "不正确" in j else "其他")
            if key not in judgments:
                judgments[key] = 0
            judgments[key] += 1

            review_lines.append(f"### {item.get('field', '')}\n")
            review_lines.append(f"- **AI提取值**: {item.get('ai_value', '')}")
            review_lines.append(f"- **参考答案**: {item.get('ref_value', '')}")
            review_lines.append(f"- **裁���结果**: **{item.get('judgment', '')}**")

            is_missing = item.get('is_field_missing', False)
            if is_missing:
                review_lines.append(f"- **结论**: ⚠️ 该字段在合同中不存在")
            else:
                review_lines.append(f"- **原文位置**: {item.get('contract_location', '未定位')}")
                review_lines.append(f"- **原文摘录**: {item.get('contract_text', '未引用')}")
            review_lines.append(f"- **复核依据**: {item.get('explanation', '')}")
            review_lines.append("")

        # 汇总
        review_lines.insert(3, f"- **判定统计**: {json.dumps(judgments, ensure_ascii=False)}\n")

        # 描述性文件名（如 20260601_四川成都_复核报告.md）
        prov = standardized.get("省", "")
        city = standardized.get("市", "")
        location = f"{prov}{city}" if prov and city else "合同"
        review_path = os.path.join(REVIEWS_DIR, f"{timestamp}_{location}_复核报告.md")
        with open(review_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(review_lines))

        # JSON存档
        json_path = os.path.join(REVIEWS_DIR, f"{timestamp}_{location}_复核报告.json")
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({"timestamp": timestamp, "discrepancies": discrepancies, "review": result,
                       "judgments": judgments}, f, ensure_ascii=False, indent=2)

        self._send_json({
            "success": True,
            "review": result,
            "review_path": review_path,
            "judgments": judgments,
            "contract_text_used": len(contract_snippet)
        })


def main():
    port = 8080
    server = HTTPServer(("0.0.0.0", port), ContractHandler)
    print(f"  ✅ 合同提取 Web 应用已启动")
    print(f"  🌐 访问地址: http://localhost:{port}")
    print(f"  📂 上传目录: {UPLOAD_DIR}")
    print(f"  📋 复盘报告: {REVIEWS_DIR}")
    print(f"  ⏹  按 Ctrl+C 停止")
    print()
    server.serve_forever()


if __name__ == "__main__":
    main()
