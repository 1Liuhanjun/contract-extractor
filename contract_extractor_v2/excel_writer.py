"""
Excel写入器
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def write_results(standardized_results, output_path, metadata_list=None):
    """
    将标准化结果写入Excel

    参数:
        standardized_results: list[dict]，每个dict是标准字段值
        output_path: 输出文件路径
        metadata_list: list[dict]，内部字段的值（登记日期/快递单号等）
    """
    from field_knowledge_base import EXCEL_HEADERS, HEADER_TO_KEY

    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    skip_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    skip_font = Font(color="999999", italic=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "合同信息提取"

    # 表头
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 数据行
    for row_offset, std_result in enumerate(standardized_results):
        row_num = row_offset + 2
        metadata = metadata_list[row_offset] if metadata_list else {}

        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            # 优先取标准化结果，其次取metadata（内部字段）
            if header in std_result and std_result[header] not in (None, "", "null"):
                value = std_result[header]
            elif header in metadata and metadata[header] not in (None, "", "null"):
                value = metadata[header]
            else:
                value = ""

            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(len(str(value)) > 20))
            cell.border = thin_border

    # 标记"不用填写"的字段（项目名称/编号 和 合同编码）
    skip_cols = [
        EXCEL_HEADERS.index("项目名称/编号") + 1,
        EXCEL_HEADERS.index("合同编码") + 1,
    ]
    for row_num in range(2, len(standardized_results) + 2):
        for col in skip_cols:
            cell = ws.cell(row=row_num, column=col)
            if cell.value in (None, "", "null"):
                cell.value = "（不用填写）"
                cell.fill = skip_fill
                cell.font = Font(color="999999", italic=True)

    # 列宽
    col_widths = {
        "登记日期": 12, "项目名称/编号": 18, "客户分类": 10, "业务类型": 10,
        "合同类型": 10, "省": 6, "市": 6, "客户名称": 35,
        "合同主体": 30, "合同编码": 22, "合同名称": 30,
        "合同开始时间": 14, "合同结束时间": 14, "是否完成签订": 12,
        "是否同步财务": 12, "合同预警提醒": 12, "账期（天）": 10,
        "保证金（万元）": 12, "税率": 10,
        "是否有旺季补偿": 12, "旺季补偿时间": 55, "旺季补偿规则": 50,
        "补偿比例": 10, "是否有油价联动": 12, "油价基准(元/升）": 12,
        "是否有疫情补贴": 12, "补贴标准": 20,
        "联系人": 10, "电话": 14, "地址": 40,
        "快递单号": 16, "钉钉审批单号": 22
    }
    for idx, header in enumerate(EXCEL_HEADERS, 1):
        width = col_widths.get(header, 12)
        ws.column_dimensions[get_column_letter(idx)].width = width

    # 行高
    ws.row_dimensions[1].height = 25
    for row_num in range(2, len(standardized_results) + 2):
        ws.row_dimensions[row_num].height = 45

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(output_path)
    count = len(standardized_results)
    print(f"\n  [完成] 已写入 {count} 条记录到: {output_path}")
    return output_path
