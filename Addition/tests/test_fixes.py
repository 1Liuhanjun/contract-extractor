"""回归测试：锁定审查发现的 P0 修复，防止回归。

- H1 缓存串档：缓存命中必须校验 sha1，不匹配则重新 OCR。
- BUG-5：小数吨位标签不被 int() 截断（2.75 ≠ 3）。
- SUSP-5：多直读价反算基准不一致时放弃换算、只留直读。
"""
import sys
import json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from layers import represent as L0
from layers import calculate as L3
from providers.ocr import OcrDoc
from providers import llm as LLM
from schemas import ContractResult, RouteRow, Coefficient, Field


def test_safe_json_tolerates_control_chars():
    """系数表 bug：字符串内含原始换行/控制字符，strict=False 应能解析。"""
    raw = '{"基准车型": "20吨\n12.5米", "系数": {"8吨": 0.63}}'
    obj = LLM._safe_json(raw)
    assert obj["系数"]["8吨"] == 0.63


def test_safe_json_salvages_truncated_routes():
    """线路明细 bug：输出超 max_tokens 被截断成残缺 JSON，应抢救出已完整的线路。"""
    truncated = ('{"found": true, "routes": ['
                 '{"线路名称":"杭州-郑州","吨位价格":{"20吨/12.5":5268}},'
                 '{"线路名称":"宁波-北京","吨位价格":{"20吨/12.5":5535}},'
                 '{"线路名称":"温州-上海","吨位价格":{"20吨/12.')  # 在第3条中途被切断
    obj = LLM._safe_json(truncated)
    assert obj["found"] is True
    assert len(obj["routes"]) == 2          # 抢救出前 2 条完整线路，丢弃残缺第3条
    assert obj["routes"][0]["线路名称"] == "杭州-郑州"


class CountingOcr:
    def __init__(self):
        self.calls = 0

    def parse_pdf(self, pdf_path, on_progress=None):
        self.calls += 1
        return OcrDoc(full_markdown="NEW", pages=["NEW"], tables=[])


def test_cache_rejects_stale_sha1(tmp_path):
    """缓存里存的是别的 sha1 → 必须忽略旧缓存重新 OCR（防同名不同内容串档）。"""
    cache = tmp_path / "c"
    cache.mkdir()
    # 预置一份"别的文件"的缓存
    (cache / "ocr.json").write_text(json.dumps({
        "sha1": "OLD_OTHER_FILE", "full_markdown": "STALE", "pages": ["STALE"], "tables": []
    }, ensure_ascii=False), encoding="utf-8")

    ocr = CountingOcr()
    # 传入与缓存不同的 sha1
    doc = L0.represent("dummy.pdf", ocr, cache, sha1="CURRENT_NEW", use_cache=True)
    assert ocr.calls == 1, "sha1 不匹配应重新 OCR"
    assert doc.full_markdown == "NEW", "不能读到旧文件的缓存内容"

    # 第二次同一 sha1 → 命中缓存，不再 OCR
    ocr2 = CountingOcr()
    doc2 = L0.represent("dummy.pdf", ocr2, cache, sha1="CURRENT_NEW", use_cache=True)
    assert ocr2.calls == 0, "sha1 一致应命中缓存"
    assert doc2.full_markdown == "NEW"


def test_lead_ton_keeps_decimal():
    assert L3._lead_ton("2.75吨/3吨") == 2.75   # 不能被截成 2
    assert L3._lead_ton("12吨/9.6米") == 12.0
    assert L3._lead_ton("20吨/12.5米") == 20.0


def _field_map():
    return {
        "tonnage_columns": [
            {"col": "L", "name": "8吨", "ton": 8},
            {"col": "O", "name": "20吨/12.5", "ton": 20},
            {"col": "P", "name": "25吨", "ton": 25},
        ],
    }


def test_city_derivation_common_origin():
    """邮政 sheet「市」取线路共同起点：起点唯一→该城市（即便省级合同）；起点跨多市→省级填'/'。"""
    fm = _field_map()
    # 市级合同，单一起点 → 该城市
    r1 = ContractResult()
    r1.coefficient = Coefficient(found=False)
    r1.routes = [RouteRow(线路名称="杭州-郑州", 吨位价格={})]
    r1.ledger_fields = {"省": Field(value="浙江"), "市": Field(value="杭州")}
    L3.compute(r1, fm)
    assert r1.routes[0].__dict__["市"] == "杭州"
    # 省级合同(台账市='/')但所有线路起点一致(成都) → 成都（修正：不再被一刀切成'/'）
    r2 = ContractResult()
    r2.coefficient = Coefficient(found=False)
    r2.routes = [RouteRow(线路名称="成都-东莞", 吨位价格={}),
                 RouteRow(线路名称="成都-北京", 吨位价格={})]
    r2.ledger_fields = {"省": Field(value="四川"), "市": Field(value="/")}
    L3.compute(r2, fm)
    assert all(r.__dict__["市"] == "成都" for r in r2.routes)
    # 省级合同 + 线路起点跨多市(杭州/宁波) → 无共同起点 → '/'
    r3 = ContractResult()
    r3.coefficient = Coefficient(found=False)
    r3.routes = [RouteRow(线路名称="杭州-郑州", 吨位价格={}),
                 RouteRow(线路名称="宁波-北京", 吨位价格={})]
    r3.ledger_fields = {"省": Field(value="浙江"), "市": Field(value="/")}
    L3.compute(r3, fm)
    assert all(r.__dict__["市"] == "/" for r in r3.routes)


def test_municipality_city_normalized_to_district():
    """直辖市合同：台账「市」被 LLM 填成'北京市'时，据客户名称里的'XX区'确定性校正为区县名；
    明细「市」随之继承该区县（修：黄村合同 市 应为'大兴'，不是'北京市'/'北京黄村'）。"""
    fm = _field_map()
    r = ContractResult()
    r.coefficient = Coefficient(found=False)
    r.routes = [RouteRow(线路名称="北京黄村-广州（京1）", 吨位价格={})]
    r.ledger_fields = {
        "省": Field(value="北京"),
        "市": Field(value="北京市", confidence=0.95),   # LLM 误填直辖市名本身
        "客户名称": Field(value="中国邮政集团有限公司北京市大兴区分公司"),
    }
    L3.compute(r, fm)
    assert r.ledger_fields["市"].value == "大兴"          # 台账市被校正为区县
    assert r.routes[0].__dict__["市"] == "大兴"           # 明细继承台账市（不取线路起点'北京黄村'）


def test_municipality_city_level_uses_municipality_name(tmp_path):
    """直辖市【市级】合同：甲方是'上海市分公司/上海市邮区中心'(名里无区县) → 市='上海'，
    绝不能用办公地址里的'浦东新区'。回归：上海合同 市 被误填'浦东新区'/'/'。"""
    fm = _field_map()
    for name, addr, llm_city in [
        ("中国邮政集团有限公司上海市邮区中心、中国邮政速递物流股份有限公司上海市分公司",
         "上海市浦东新区龙东大道4877号", "浦东新区"),   # LLM 从地址误抠区县
        ("中国邮政集团有限公司上海市分公司", "上海市浦东新区龙东大道4877号", "/"),   # 旧版误填'/'
    ]:
        r = ContractResult()
        r.coefficient = Coefficient(found=False)
        r.routes = [RouteRow(线路名称="上海-福州", 吨位价格={})]
        r.ledger_fields = {"省": Field(value="上海"),
                           "市": Field(value=llm_city, confidence=0.5),
                           "客户名称": Field(value=name),
                           "地址": Field(value=addr)}
        L3.compute(r, fm)
        assert r.ledger_fields["市"].value == "上海", (name, r.ledger_fields["市"].value)
        assert r.routes[0].__dict__["市"] == "上海"     # 明细继承台账市=上海


def test_city_district_extraction_no_overcapture():
    """归一正则不应跨字误捕（'北京市大兴区分公司' → '大兴'，而非'京市大兴'之类）。"""
    fm = _field_map()
    for name, expect in [
        ("中国邮政集团有限公司北京市大兴区分公司", "大兴"),
        ("中国邮政集团有限公司北京市石景山区分公司", "石景山"),
        ("中国邮政集团有限公司上海市浦东新区分公司", "浦东新"),  # '区'前最短匹配
    ]:
        r = ContractResult()
        r.coefficient = Coefficient(found=False)
        r.routes = [RouteRow(线路名称="甲-乙", 吨位价格={})]
        r.ledger_fields = {"省": Field(value="北京"),
                           "市": Field(value=None),
                           "客户名称": Field(value=name)}
        L3.compute(r, fm)
        assert r.ledger_fields["市"].value == expect, (name, r.ledger_fields["市"].value)


def test_municipality_normalize_leaves_real_city_untouched():
    """非直辖市/已是具体市的台账「市」不被归一逻辑改动（南京合同保持南京，且明细继承南京不取起点）。"""
    fm = _field_map()
    r = ContractResult()
    r.coefficient = Coefficient(found=False)
    r.routes = [RouteRow(线路名称="徐州-福州", 吨位价格={})]   # 起点徐州 ≠ 甲方城市南京
    r.ledger_fields = {"省": Field(value="江苏"),
                       "市": Field(value="南京"),
                       "客户名称": Field(value="中国邮政集团有限公司江苏省南京市分公司")}
    L3.compute(r, fm)
    assert r.ledger_fields["市"].value == "南京"
    assert r.routes[0].__dict__["市"] == "南京"            # 继承台账市，不被填成起点'徐州'


def test_no_business_field_uses_fixed_default():
    """防回归：业务字段一律不得使用固定默认值（config.defaults 不得含业务字段；
    field_map 不得有 source:D 列）。合同主体/公司主体/客户分类必须改为抽取/继承/注入。"""
    from app_config import Config
    cfg = Config()
    # config.defaults 不得再含会"自信填错"的业务默认值
    for bad in ("合同主体", "公司主体", "客户分类"):
        assert bad not in (cfg.defaults or {}), f"{bad} 不应再作为固定默认值"
    # field_map 任何列都不应再用 source:D（固定默认值）
    for seg in ("ledger", "youzheng"):
        d_cols = [c["name"] for c in cfg.field_map[seg] if c.get("source") == "D"]
        assert not d_cols, f"{seg} 仍有 source:D 固定默认值列: {d_cols}"
    # 合同主体/公司主体改为抽取/继承
    led = {c["name"]: c for c in cfg.field_map["ledger"]}
    yz = {c["name"]: c for c in cfg.field_map["youzheng"]}
    assert led["合同主体"]["source"] == "A" and led["合同主体"].get("field") == "合同主体"
    assert yz["公司主体"].get("from") == "ledger" and yz["公司主体"].get("field") == "合同主体"
    assert led["客户分类"].get("no_llm") is True


def test_extension_routes_detected():
    """含『比照/延申/新增邮路』机制 → 给出人工补充提示；普通合同 → 不提示。"""
    from layers import validate as L4
    md = ("主要中标价格明细表（直读13条）……2、在起点和终点确定的情况下，若将来出现串行、"
          "环形、相似等新增邮路需求，则先比照……计算公式为：单程邮路协议价（元/趟）÷原邮路载重"
          "吨位÷单程邮路里程×新邮路载重吨位×串行邮路里程。")
    w = L4.check_extension_routes(md, 13)
    assert w and "比照/延申" in w and "13 条" in w
    assert "计算公式" not in w or "协议价" in w  # 折算公式被摘出
    # 普通合同无该机制 → None
    assert L4.check_extension_routes("普通合同，价格表已列全部线路。", 5) is None


def test_route_nature_out_of_enum_flagged():
    """邮路性质枚举外（如"单边"）：保留原文，但触发复核告警（不自动归一）。"""
    from layers import validate as L4
    result = ContractResult()
    result.coefficient = Coefficient(found=False)
    result.routes = [RouteRow(线路名称="甲-乙", 邮路性质="单边", confidence=0.95),
                     RouteRow(线路名称="丙-丁", 邮路性质="双程", confidence=0.95)]
    result.ledger_fields = {"省": Field(value="浙江")}
    warnings = L4.validate(result, _field_map())
    # "单边"非标准枚举 → 有告警；"双程"是标准枚举 → 无告警
    nat_warns = [w for w in warnings if "邮路性质" in w]
    assert len(nat_warns) == 1 and "单边" in nat_warns[0]
    # 原文保留，未被改成"单程"
    assert result.routes[0].邮路性质 == "单边"


def test_uncertain_cells_highlighted(tmp_path):
    """不确定字段在生成的 Excel 中应黄色高亮 + 批注；确定字段不高亮。"""
    from app_config import Config
    from layers import excel_writer as L5
    import openpyxl
    cfg = Config()
    out = tmp_path / "t.xlsx"
    ledger = {
        "账期": Field(value=30, confidence=0.55, evidence_ok=False),       # 不确定
        "客户名称": Field(value="中国邮政XX", confidence=1.0, evidence_ok=True),  # 确定
    }
    L5.write_contract(cfg.template_xlsx, out, ledger, [], cfg.field_map, cfg.defaults)
    ws = next(w for w in openpyxl.load_workbook(out).worksheets
              if w.title.startswith("新合同台账"))
    q = ws["Q2"]   # 账期（天）→ 不确定
    assert q.fill.patternType == "solid" and str(q.fill.fgColor.rgb).endswith("FFF3CD")
    assert q.comment and "核对" in q.comment.text
    assert ws["H2"].fill.patternType != "solid"   # 客户名称确定 → 不高亮


def test_price_cells_padded_to_two_decimals(tmp_path):
    """价格列(吨位价)写 Excel 时补齐到至少 2 位小数：整数 111→显示 111.00、值仍为数值。
    2 位及以上(5307.12)原样保留。"""
    from app_config import Config
    from layers import excel_writer as L5
    import openpyxl
    cfg = Config()
    out = tmp_path / "p.xlsx"
    route = RouteRow(线路名称="甲-乙", 邮路性质="往返", 里程=100)
    route.__dict__["最终吨位价格"] = {"5吨": 111.0, "8吨": 5307.12}  # J列=5吨, L列=8吨
    ledger = {"省": Field(value="江苏"), "市": Field(value="南京")}
    L5.write_contract(cfg.template_xlsx, out, ledger, [route], cfg.field_map, cfg.defaults)
    ws = next(w for w in openpyxl.load_workbook(out).worksheets
              if w.title.startswith(cfg.field_map["youzheng_sheet"]))
    j, l = ws["J2"], ws["L2"]
    assert j.value == 111.0 and j.number_format == "0.00"        # 整数 → 数值 + 2 位格式
    assert l.value == 5307.12 and l.number_format == "0.00"      # 2 位价原样、格式一致


def test_inconsistent_base_abandons_conversion():
    """两个直读价反算基准不一致(>0.5%) → 不换算，只保留直读。"""
    fm = _field_map()
    result = ContractResult()
    result.coefficient = Coefficient(found=True, 基准车型="20吨/12.5米",
                                     系数={"8吨": 0.63, "25吨": 1.14})
    # 直读 8吨=100（基准≈158.7）与 25吨=200（基准≈175.4）反算基准差 >0.5%
    result.routes = [RouteRow(线路名称="甲-乙", 吨位价格={"8吨": 100.0, "25吨": 200.0})]
    result.ledger_fields = {"省": Field(value="浙江")}
    notes = L3.compute(result, fm)

    final = result.routes[0].__dict__["最终吨位价格"]
    # 只应保留直读的 8吨/25吨，不应换算出 20吨
    assert set(final.keys()) == {"8吨", "25吨"}
    assert "20吨/12.5" not in final
    assert any("放弃换算" in (result.routes[0].换算说明 or "") for _ in [0])
    assert any("不一致" in n for n in notes)
