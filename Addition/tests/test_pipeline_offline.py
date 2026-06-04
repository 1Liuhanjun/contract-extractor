"""离线端到端验证：用 Mock provider 跑通完整 pipeline，校验关键产出。

不需要任何 API key。验证点对应需求文档：
- §4 流程跑通（分类→抽取→换算→写表→报告）
- §6.1 跨吨位换算数值正确（基准 20 吨 → 8/12/25/30）
- §3.4/§6.3 Excel 结构与 P 列公式
- §11 重复运行去重
"""
import sys
import shutil
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string

CODE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(CODE_DIR))

from app_config import Config
from providers.mock import MockLlm, MockOcr
from layers import excel_writer as L5
import pipeline


SAMPLE_PDF = CODE_DIR / "samples" / "浙江.pdf"


@pytest.fixture
def env(tmp_path):
    cfg = Config()
    cfg.output_dir = tmp_path / "output"
    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    return cfg


def _out(cfg, name):
    return cfg.output_dir / "contracts" / f"{name}.xlsx"


def _cell(ws, col, row):
    return ws.cell(row=row, column=column_index_from_string(col)).value


def test_full_pipeline(env):
    cfg = env
    assert SAMPLE_PDF.exists(), f"缺少样例 PDF：{SAMPLE_PDF}"
    llm, ocr = MockLlm(), MockOcr()
    index = {}

    out = _out(cfg, "浙江")
    r = pipeline.process_pdf(SAMPLE_PDF, cfg, llm, ocr, out, index)
    assert r["status"] == "ok", r
    assert r["rows"] == 2, "应写入 2 条线路"
    assert r["xlsx_name"] == out.name

    wb = openpyxl.load_workbook(out)
    ledger = next(w for w in wb.worksheets if w.title.startswith("新合同台账"))
    yz = wb["邮政"]

    # —— 台账行（第2行）——
    assert _cell(ledger, "C", 2) == "邮政"                       # 客户分类：分类结果注入(no_llm)，非固定默认值
    assert _cell(ledger, "I", 2) == "北京博华物流有限公司"        # 合同主体：抽自乙方(本样例乙方=博华)，非固定默认值
    assert str(_cell(ledger, "Q", 2)) == "60"                    # 账期
    # 红字列留空（同事口径：B项目编号、A登记日期、K合同名称 等不抽取）
    assert _cell(ledger, "B", 2) in (None, "")
    assert _cell(ledger, "A", 2) in (None, "")          # 登记日期→红字留空
    assert _cell(ledger, "K", 2) in (None, "")          # 合同名称→红字留空
    # 是否完成签订（N 列）改为可抽字段（对齐同事），应写入"是"
    assert _cell(ledger, "N", 2) == "是"
    # P 列保留公式
    assert str(_cell(ledger, "P", 2)).startswith("=IFERROR(DATEDIF")

    # —— 邮政明细：换算校验（基准 20 吨=100 → 8/12/25/30）——
    assert _cell(yz, "F", 2) == "杭州-郑州"
    assert _cell(yz, "B", 2) == "/"                              # 省级合同(台账市="/") → 市列填"/"（§4.2）
    assert _cell(yz, "O", 2) == 100.0                            # 20吨直读
    assert _cell(yz, "L", 2) == 63.0                             # 8吨=100*0.63
    assert _cell(yz, "M", 2) == 74.0                             # 12吨=100*0.74
    assert _cell(yz, "P", 2) == 114.0                            # 25吨=100*1.14
    assert _cell(yz, "Q", 2) == 127.0                            # 30吨=100*1.27
    assert _cell(yz, "E", 2) == "1"                              # 分包号（红字例外，要抽）
    assert _cell(yz, "V", 2) == "北京博华物流有限公司"            # 公司主体：继承台账「合同主体」(乙方)，非固定值
    # 红字不抽取列：即便 mock LLM 在 _routes 里返回了 类型/合同编号，也不应写入 Excel（留空）
    assert _cell(yz, "AA", 2) in (None, "")                      # 合同编号（红字不抽取）
    assert _cell(yz, "W", 2) in (None, "")                       # 类型（红字不抽取）
    assert _cell(yz, "X", 2) in (None, "")                       # 备注（红字不抽取）
    assert _cell(yz, "S", 2) in (None, "")                       # 元/趟/条/40吨B（红字不抽取）

    # 第二条线路按 200 换算
    assert _cell(yz, "O", 3) == 200.0
    assert _cell(yz, "L", 3) == 126.0                            # 200*0.63

    # —— 复核报告生成且含保证金留空 ——
    report = Path(r["report"])
    assert report.exists()
    txt = report.read_text(encoding="utf-8")
    assert "保证金" in txt and "留空" in txt


def test_reprocess_allows_reupload(env):
    """已取消二次上传跳过：重复 PDF 应重新处理(ok + reprocessed)，且各自生成独立台账、互不追加。"""
    cfg = env
    llm, ocr = MockLlm(), MockOcr()
    index = {}
    out1 = _out(cfg, "job1_浙江")
    out2 = _out(cfg, "job2_浙江")
    r1 = pipeline.process_pdf(SAMPLE_PDF, cfg, llm, ocr, out1, index)
    r2 = pipeline.process_pdf(SAMPLE_PDF, cfg, llm, ocr, out2, index)
    assert r1["status"] == "ok"
    assert r2["status"] == "ok", "重传应重新处理，不再跳过"
    assert r2.get("reprocessed") is True, "重传结果应标记 reprocessed=True"
    # 两次各自独立成文件，各自只有 2 行（表头+2）→ 证明是独立新建、互不追加
    for out in (out1, out2):
        wb = openpyxl.load_workbook(out)
        assert wb["邮政"].max_row == 3, f"{out.name} 应表头+2行，实际 {wb['邮政'].max_row}"


def test_unknown_company(env, tmp_path):
    cfg = env

    class BlankOcr(MockOcr):
        def parse_pdf(self, pdf_path, on_progress=None):
            doc = super().parse_pdf(pdf_path)
            doc.full_markdown = "某无关公司服务协议，无任何客户关键词。"
            return doc

    r = pipeline.process_pdf(SAMPLE_PDF, cfg, MockLlm(), BlankOcr(), _out(cfg, "x"), {})
    assert r["status"] == "skipped_未知"
    assert (cfg.output_dir / "_unknown").exists()
