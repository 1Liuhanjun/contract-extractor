"""第5层 写台账 Excel。

每份合同写成一份**独立**的台账文件（write_contract）：从模板复制结构、清空数据行，
只填这一份合同的 1 行台账 + N 行明细，互不共用文件（彻底隔离，避免跨合同污染）。

- 仅 P 列保留 DATEDIF 公式，其余写数值/文本。
- 日期列写 datetime 并设 number_format。
- index（output/_index.json）仅用于历史列表 / 软去重告警 / 重传标记，不再控制写入位置。
"""
import json
import shutil
import logging
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

from schemas import ENUM_ROUTE_NATURE

warnings.filterwarnings("ignore")
log = logging.getLogger("excel")

# 不确定数据高亮：柔和琥珀黄填充 + 批注说明原因（与复核报告"不确定字段"口径一致）
LOW_CONF = 0.7
_UNCERTAIN_FILL = PatternFill("solid", fgColor="FFF3CD")


def _uncertain_reason(f):
    """抽取字段是否"不确定"：有多候选 / 证据未命中 / 置信偏低。返回原因串或 None。"""
    if f is None or getattr(f, "value", None) in (None, "", "null"):
        return None
    rs = []
    if f.candidates and len(f.candidates) > 1:
        rs.append("多次抽取有分歧")
    if f.evidence_ok is False:
        rs.append("证据未在原文命中")
    if (f.confidence or 1) < LOW_CONF:
        rs.append(f"置信偏低({f.confidence})")
    return "；".join(rs) if rs else None


def _mark_uncertain(ws, col_letter, row, reason):
    """给单元格加黄色高亮 + 批注（仅当该格已写入值时调用）。"""
    cell = ws.cell(row=row, column=column_index_from_string(col_letter))
    if cell.value in (None, ""):
        return
    cell.fill = _UNCERTAIN_FILL
    if reason:
        cell.comment = Comment(f"⚠ 待人工核对：{reason}", "合同读取智能体")


def _find_sheet(wb, prefix):
    for ws in wb.worksheets:
        if ws.title.startswith(prefix):
            return ws
    return None


def _next_row(ws):
    """下一可写行（表头占第1行）。"""
    r = ws.max_row
    # max_row 在空表可能仍返回 1；逐列确认第1行后是否真为空
    if r == 1:
        return 2
    return r + 1


def _to_date(v):
    if isinstance(v, datetime):
        return v
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _set_cell(ws, col_letter, row, value, kind=None):
    if value is None or value == "":
        return
    cell = ws.cell(row=row, column=column_index_from_string(col_letter))
    if kind == "date":
        d = _to_date(value)
        if d is not None:
            cell.value = d
            cell.number_format = "yyyy-mm-dd"
            return
    if kind == "price":
        # 价格列：写成数值并强制至少 2 位小数显示（111→111.00；5307.12 原样）。
        # 仍是 number（非文本），可正常求和/排序；值本身 ≤2 位小数（换算用 round(_,2)），故不会被截断。
        try:
            cell.value = float(value)
            cell.number_format = "0.00"
            return
        except (TypeError, ValueError):
            pass  # 非数值（极少数 OCR 残留字符串）→ 按原值落，不强格式
    cell.value = value


def _ledger_value(col, ledger_fields, defaults):
    src = col.get("source")
    if col.get("extract") is False:
        return defaults.get(col["name"])  # 红字：留空，除非配置了默认
    if src == "D":
        return defaults.get(col.get("default"))
    if src in ("A", "B"):
        f = ledger_fields.get(col.get("field", col["name"]))
        return f.value if f else None
    return None


def write_ledger_row(master_path: Path, ledger_fields, field_map, defaults) -> int:
    wb = openpyxl.load_workbook(master_path)
    ws = _find_sheet(wb, field_map["ledger_sheet_prefix"])
    row = _next_row(ws)
    for col in field_map["ledger"]:
        if col.get("formula"):
            _set_cell(ws, col["col"], row, col["formula"].replace("{row}", str(row)))
            continue
        val = _ledger_value(col, ledger_fields, defaults)
        _set_cell(ws, col["col"], row, val, kind=col.get("kind"))
        # 高亮不确定的抽取字段（source A 直读字段）
        if col.get("source") == "A" and col.get("extract", True) and not col.get("from"):
            reason = _uncertain_reason(ledger_fields.get(col.get("field", col["name"])))
            if reason:
                _mark_uncertain(ws, col["col"], row, reason)
    wb.save(master_path)
    return row


def _youzheng_value(col, route, ledger_fields, defaults):
    if col.get("extract") is False:
        return defaults.get(col["name"])
    src = col.get("source")
    if src == "D":
        return defaults.get(col.get("default"))
    if col.get("tonnage"):
        return (route.__dict__.get("最终吨位价格", {}) or {}).get(col["name"])
    # 来自台账的派生字段
    if col.get("from") == "ledger":
        field = col.get("field", col["name"])
        if field in ("省", "业务类型"):
            f = ledger_fields.get(field)
            return f.value if f else None
        f = ledger_fields.get(field)
        return f.value if f else None
    # 派生省/市
    if col["name"] == "省":
        return route.__dict__.get("省")
    if col["name"] == "市":
        return route.__dict__.get("市")
    # 直接来自 route 的属性
    field = col.get("field", col["name"])
    return getattr(route, field, None)


def write_contract(template_path: Path, out_path: Path, ledger_fields, routes,
                   field_map, defaults) -> int:
    """为单份合同生成**独立**台账文件并填充，返回写入的邮政明细行数。

    每次调用都从模板新建/覆盖 out_path（复制结构+清空数据行），只写这一份合同：
    台账 sheet 第 2 行、邮政 sheet 从第 2 行起。绝不与其它合同共用文件。
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    for ws in wb.worksheets:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    wb.save(out_path)
    # 空表 → 既有写入逻辑会从第 2 行开始写
    write_ledger_row(out_path, ledger_fields, field_map, defaults)
    return write_youzheng_rows(out_path, routes, ledger_fields, field_map, defaults)


def _youzheng_uncertain(col, route, ledger_fields):
    """邮政明细某列是否"不确定"。返回原因串或 None。"""
    # 取自台账的字段（业务类型/省/有效期起止…）→ 跟随该台账字段的不确定性
    if col.get("from") == "ledger":
        return _uncertain_reason(ledger_fields.get(col.get("field", col["name"])))
    # 线路名称：整条线路置信低（线路名未在原文命中）→ 标记该行关键单元格
    if col.get("name") == "线路名称" and (route.confidence or 1) < LOW_CONF:
        return "线路名未在原文命中，请核对本行整条线路"
    # 邮路性质：枚举外的值（如"单边/双边"）保留原文但标低置信、提示人工确认
    if col.get("name") == "邮路性质":
        nat = route.邮路性质
        if nat and str(nat).strip() and str(nat).strip() not in ENUM_ROUTE_NATURE:
            return f"邮路性质「{nat}」非标准枚举（{'/'.join(ENUM_ROUTE_NATURE)}），请人工确认"
    return None


def write_youzheng_rows(master_path: Path, routes, ledger_fields, field_map, defaults) -> int:
    wb = openpyxl.load_workbook(master_path)
    ws = _find_sheet(wb, field_map["youzheng_sheet"])
    start = _next_row(ws)
    row = start
    for route in routes:
        for col in field_map["youzheng"]:
            val = _youzheng_value(col, route, ledger_fields, defaults)
            kind = "price" if col.get("tonnage") else col.get("kind")
            _set_cell(ws, col["col"], row, val, kind=kind)
            reason = _youzheng_uncertain(col, route, ledger_fields)
            if reason:
                _mark_uncertain(ws, col["col"], row, reason)
        row += 1
    wb.save(master_path)
    return row - start  # 写入行数


# ---- 去重 ----
def load_index(output_dir: Path) -> dict:
    p = output_dir / "_index.json"
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return {}


def save_index(output_dir: Path, index: dict):
    (output_dir / "_index.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=1), encoding="utf-8")


def is_duplicate(index: dict, sha1: str) -> bool:
    return sha1 in index
